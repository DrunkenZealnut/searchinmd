#!/usr/bin/env python3
"""
안전등급 재채점 — 원본 페이지 본문에서 등급을 다시 계산한다.

`recount_grades.py` 는 원본 워크북의 **기존 등급을 재매핑**할 뿐이다. 이 스크립트는
`페이지전체내용` 열에서 등급을 **처음부터 다시 계산**한다.

    pip install openpyxl
    python3 regrade.py --validate        # 현재 규칙 재현율만 확인 (산출물 안 씀)
    python3 regrade.py                   # 재채점 + 항목별 영향도 표

## 현재 규칙 (커밋된 등급사유 4,000여 건에서 역설계)

    안전어 ≤ 5                → 등급1  (1,567쪽에서 100% 일관)
    안전어 > 5, 조치어 ≤ 4    → 등급2
    안전어 > 5, 조치어 ≥ 5    → 등급3  (431쪽에서 겹침 0으로 분리)

## 고치는 결함 셋

D1 단어 경계 없음 — `진동`·`먼지` 같은 반도체 문맥 동음이의가 안전어로 잡힌다.
   실측 175쪽이 이 넷만으로 등급이 정해졌다. 또 `안전` 이 `산업안전보건법`·
   `안전장치` 안에도 들어 있어 한 출현이 두 번 세어진다.
D2 길이 정규화 없음 — 임계 5건이 페이지 길이와 무관하다. 페이지 길이 편차가
   커서(43%가 1,000자 미만, 최대 32,767자) 긴 페이지가 그냥 승급한다.
D3 총계 off-by-one — 3쪽(0.16%). 나머지 168쪽의 "불일치" 는 사유 문자열이
   상위 5개만 보여주는 표시 절단이지 채점 버그가 아니다.

각 결함을 따로 켜고 끌 수 있어 영향도를 항목별로 뗄 수 있다.
"""
import argparse
import json
import os
import re
import sys
from collections import Counter, defaultdict

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit('openpyxl이 필요합니다: pip install openpyxl')

HERE = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(HERE, 'data')
OUT_DIR = os.path.join(HERE, 'docs', '03-analysis', 'data')

NCS_FILE = 'ncs_keywords_in_markdown_results_20260402.xlsx'
TXT_FILE = 'ncs_keywords_in_markdown_results_교과서_results_20260415.xlsx'

# 등급사유에 실제로 등장한 어휘 전부. 손으로 고른 목록이 아니라 커밋된 CSV 의
# 4,000여 건에서 뽑았다 — 원본 채점기가 쓰던 사전 그 자체다.
SAFETY_TERMS = [
    '안전', '위험', '주의', '유의 사항', '유의사항', '사고', '유해', '진동', '누출',
    'MSDS', '보건', '폭발', '화재', '먼지', '화학물질', '보호구', '소음', '방사선',
    '감전', '질식', '분진', '중독', '끼임', '작업환경', '물질안전보건자료', '인화',
    '부상', '산업안전보건법', '질병', 'PSM', '추락', '사망',
]
ACTION_TERMS = [
    '방지', '예방', '착용', '환기', '차단', '안전관리', '대처', '마스크', '안전장치',
    '안전화', '대피', '응급조치', '격리', '소화기', '보안경', '보호복', '안전모',
    '안전교육', '방독면', '안전수칙', '조치사항', '보호장비', '안전대책', '안전조치',
    '관리방법',
]

SAFETY_MIN = 6          # 이 이상이어야 등급1을 벗어난다 (원본의 "5건 이하" 경계)
ACTION_MIN = 5          # 이 이상이면 등급3
DENSITY_BASE = 1000     # D2: 1,000자당으로 환산

# D1: 반도체 문맥에서 안전과 무관하게 쓰이는 출현. 용어 자체를 버리지 않고
# 이 문맥에서 나온 것만 뺀다 — 진동·소음은 실제 직업병 인자이기도 하다.
HOMONYM_CONTEXTS = {
    '진동': ['진동자', '진동수', '진동식', '초음파 진동', '진동 주파수', '격자 진동',
             '분자 진동', '진동 모드', '진동 스펙트럼'],
    '먼지': ['파티클', '미세 입자', '먼지 입자 수', '파티클 카운트'],
    '분진': ['분진 입자 수'],
    '소음': ['신호 대 잡음', '노이즈 비', '소음 지수'],
}

# D1: 더 긴 용어의 부분 문자열로 들어가 두 번 세어지는 것들.
# 예) '산업안전보건법' 한 번 나오면 naive 매칭은 '안전' 도 한 번 더 센다.
CONTAINING = {
    '안전': ['산업안전보건법', '물질안전보건자료', '안전관리', '안전장치', '안전화',
             '안전모', '안전교육', '안전수칙', '안전대책', '안전조치', '안전보건'],
    '보건': ['산업안전보건법', '물질안전보건자료', '안전보건'],
}


def count_terms(text, terms, word_boundary=False):
    """용어별 출현 수. word_boundary=True 면 D1 보정을 적용한다."""
    out = Counter()
    for t in terms:
        n = text.count(t)
        if not n:
            continue
        if word_boundary:
            # 더 긴 용어에 흡수된 출현을 뺀다
            for longer in CONTAINING.get(t, []):
                if longer != t:
                    n -= text.count(longer)
            # 반도체 문맥의 동음이의 출현을 뺀다
            for ctx in HOMONYM_CONTEXTS.get(t, []):
                n -= text.count(ctx)
            n = max(n, 0)
        if n:
            out[t] = n
    return out


def length_scale(text_len, base):
    """D2 임계 배율.

    제곱근을 쓴다. 선형(len/base)은 "안전 내용이 페이지 길이에 비례해야 한다" 는
    가정인데 근거가 없다 — 32,767자 페이지에 32배의 안전어를 요구하게 된다.
    제곱근은 긴 페이지에 불이익을 주되 그 강도를 완만하게 한다.

    base 는 임의 상수가 아니라 **해당 데이터의 페이지 길이 중앙값**이다. 중앙값
    길이의 페이지는 배율 1.0 이 되어 원본 임계(6건/5건)가 그대로 적용된다.
    원본 임계가 암묵적으로 '보통 길이 페이지' 기준으로 잡혔다고 보는 것이다.
    base 미만 페이지는 배율을 1.0 으로 묶어 짧다는 이유로 승급하지 않게 한다.
    """
    import math
    return math.sqrt(max(text_len, base) / base)


def grade_page(text, word_boundary=False, normalize=False):
    """페이지 본문 하나에 등급을 매긴다. 반환: (등급, 안전수, 조치수, 사유)

    normalize 는 False 이거나 length_scale() 이 돌려준 배율(float)이다.
    """
    s = count_terms(text, SAFETY_TERMS, word_boundary)
    a = count_terms(text, ACTION_TERMS, word_boundary)
    sn, an = sum(s.values()), sum(a.values())

    s_min, a_min = SAFETY_MIN, ACTION_MIN
    if normalize:
        s_min, a_min = SAFETY_MIN * normalize, ACTION_MIN * normalize

    if sn < s_min:
        g = 1
    elif an >= a_min:
        g = 3
    else:
        g = 2

    top = ', '.join('%s(%d)' % (k, v) for k, v in s.most_common(5))
    reason = '안전 %d건 [%s]' % (sn, top)
    if an:
        reason += ', 조치 %d건 [%s]' % (an, ', '.join(
            '%s(%d)' % (k, v) for k, v in a.most_common(5)))
    else:
        reason += ', 구체적 조치 언급 없음'
    return g, sn, an, reason


# 열은 위치로 읽는다. 헤더 텍스트가 시트마다 다르기 때문이다 — '사망' 시트는
# 헤더 행이 아예 없고, '부상'·'끼임' 등은 4번 열을 '페이지전체내용' 으로 잘못
# 이름 붙여 놓았다(실제로는 page). 30개 시트 전부 9열이고 위치는 일정하다(실측).
# recount_grades.py 가 헤더 대신 앵커 탐색을 쓰는 것도 같은 이유다.
COL_FILENAME, COL_PAGE, COL_TEXT, COL_GRADE = 2, 4, 5, 7
N_COLS = 9


def load_pages(path):
    """워크북에서 고유 (교재, 페이지) -> {text, grade} 를 모은다.

    시트는 키워드별이고 같은 페이지가 여러 시트에 중복 등장한다. 등급은 페이지
    속성이므로 먼저 고유 페이지로 접는다.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    pages, skipped = {}, 0
    for sn in wb.sheetnames:
        for row in wb[sn].iter_rows(values_only=True):
            if not row or len(row) < N_COLS:
                continue
            if str(row[0]).strip() == 'number':          # 헤더 행
                continue
            fn, pg, tx, gr = (row[COL_FILENAME], row[COL_PAGE],
                              row[COL_TEXT], row[COL_GRADE])
            if not fn or pg is None or not isinstance(tx, str) or not tx.strip():
                skipped += 1
                continue
            try:
                pg = int(float(str(pg)))
            except (TypeError, ValueError):
                skipped += 1
                continue
            key = (str(fn), pg)
            if key not in pages:
                pages[key] = {'text': tx, 'grade': gr}
    wb.close()
    if skipped:
        print('  (형식이 맞지 않아 건너뛴 행 %d개)' % skipped)
    return pages


def median_length(pages):
    L = sorted(len(r['text']) for r in pages.values())
    return L[len(L) // 2] if L else DENSITY_BASE


def run(pages, word_boundary, normalize, base=None):
    """normalize=True 면 페이지별 길이 배율을 적용한다."""
    out = {}
    for key, rec in pages.items():
        scale = length_scale(len(rec['text']), base) if normalize else False
        g, sn, an, reason = grade_page(rec['text'], word_boundary, scale)
        out[key] = {'g': g, 'sn': sn, 'an': an, 'reason': reason,
                    'len': len(rec['text'])}
    return out


def dist(graded):
    c = Counter(v['g'] for v in graded.values())
    return {g: c.get(g, 0) for g in (1, 2, 3)}


def agree(graded, pages):
    """원본 등급과 일치하는 비율."""
    ok = tot = 0
    for key, rec in pages.items():
        try:
            orig = int(rec['grade'])
        except (TypeError, ValueError):
            continue
        tot += 1
        ok += (graded[key]['g'] == orig)
    return ok, tot


def main():
    ap = argparse.ArgumentParser(description='안전등급 재채점')
    ap.add_argument('--data', default=DATA_DIR, help='원본 엑셀 위치')
    ap.add_argument('--validate', action='store_true',
                    help='현재 규칙 재현율만 출력하고 산출물은 쓰지 않는다')
    args = ap.parse_args()

    src = os.path.join(args.data, NCS_FILE)
    if not os.path.exists(src):
        sys.exit('원본 엑셀을 찾을 수 없습니다: %s\n'
                 '  --data 로 위치를 지정하세요.' % src)

    print('원본 읽는 중… %s' % os.path.basename(src))
    pages = load_pages(src)
    print('  고유 페이지 %d쪽' % len(pages))

    base = run(pages, word_boundary=False, normalize=False)
    ok, tot = agree(base, pages)
    print('\n=== 현재 규칙 재현 ===')
    print('  원본 등급과 일치: %d/%d (%.1f%%)' % (ok, tot, ok * 100.0 / max(tot, 1)))
    print('  분포: %s' % dist(base))

    if args.validate:
        return

    med = median_length(pages)
    print('  페이지 길이 중앙값 %d자 (D2 기준 길이로 사용)' % med)

    variants = [
        ('D1 단어 경계', True, False),
        ('D2 길이 정규화', False, True),
        ('D1+D2 둘 다', True, True),
    ]
    print('\n=== 항목별 영향도 (재현 기준 대비) ===')
    results = {}
    for name, wb_, nm in variants:
        g = run(pages, wb_, nm, med)
        results[name] = g
        moved = sum(1 for k in g if g[k]['g'] != base[k]['g'])
        d = dist(g)
        print('  %-16s 분포 %s  등급3 %.1f%%  변동 %d쪽 (%.1f%%)'
              % (name, d, d[3] * 100.0 / len(g), moved, moved * 100.0 / max(len(g), 1)))

    final = results['D1+D2 둘 다']
    print('\n=== 최종 (D1+D2) 대비 원본 ===')
    move = Counter()
    for k in final:
        try:
            o = int(pages[k]['grade'])
        except (TypeError, ValueError):
            continue
        if o != final[k]['g']:
            move['%d→%d' % (o, final[k]['g'])] += 1
    for k, v in sorted(move.items()):
        print('  등급 %s : %d쪽' % (k, v))

    os.makedirs(OUT_DIR, exist_ok=True)
    out = os.path.join(OUT_DIR, 'regrade_impact.json')
    payload = {
        'source': os.path.basename(src),
        'pages': len(pages),
        'median_page_len': med,
        'rule': {'safety_min': SAFETY_MIN, 'action_min': ACTION_MIN,
                 'normalize': 'sqrt(len/median)'},
        'reproduction': {'agree': ok, 'total': tot,
                         'rate': round(ok * 100.0 / max(tot, 1), 2)},
        'dist': {'baseline': dist(base),
                 **{n: dist(g) for n, g in results.items()}},
        'moves_vs_original': dict(move),
    }
    tmp = out + '.tmp'
    with open(tmp, 'w', encoding='utf-8') as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    os.replace(tmp, out)
    print('\n영향도 요약을 %s 에 썼습니다.' % os.path.relpath(out, HERE))
    print('대시보드 수치는 아직 바꾸지 않았습니다 — 반영 여부는 별도 판단입니다.')


if __name__ == '__main__':
    main()
