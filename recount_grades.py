#!/usr/bin/env python3
"""
data/ 원본 엑셀 → 통일 등급체계로 전수 재집계.

배경
----
NCS(7,769행)와 교과서(981행)는 원본 엑셀의 등급 코딩이 서로 반대다.
등급사유 문자열을 전수 분류해 확인한 결과:

  NCS  (ncs_keywords_in_markdown_results_20260402.xlsx)
    1 = "키워드 N건 (5건 이하)"               → 미흡·없음   (2,076행 100%)
    2 = "안전 N건 [...], 구체적 조치 언급 없음"  → 형식적 언급 (3,465행 100%)
    3 = "안전 N건 [...] + 조치 M건 [...]"       → 구체적 대책 (2,228행 100%)

  교과서 (ncs_keywords_in_markdown_results_교과서_results_20260415.xlsx)
    1 = "... (구체적 대책 미흡)"                → 형식적 언급 (360행)
    2 = "... (구체적 대책 포함)"                → 구체적 대책 (64행)
    3 = "키워드 N건 (5건 이하)"                → 미흡·없음   (557행 100%)

따라서 교과서는 {1→2, 2→3, 3→1}로 재매핑해 통일 등급체계
(1 미흡·없음 / 2 형식적 언급 / 3 구체적 대책)로 맞춘다.

집계 단위
--------
등급은 사실상 **페이지 단위** 판정이다(NCS 1,847쪽 중 등급이 갈린 페이지 12쪽 = 0.6%,
교과서 362쪽 중 0쪽). 행 단위 집계는 한 페이지에 걸린 키워드 수만큼 중복 계수되어
상위 등급을 체계적으로 과대평가하므로(NCS 등급3은 쪽당 평균 20.6개 키워드 → 20배 증폭),
주 지표는 고유 페이지 기준으로 산출한다.

사용법
-----
    pip install openpyxl
    python3 recount_grades.py [--out docs/03-analysis/data]
"""
import argparse
import collections
import csv
import hashlib
import json
import os
import re
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit('openpyxl이 필요합니다:  pip install openpyxl')

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from page_utils import is_cell_truncated  # noqa: E402

_HERE = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(_HERE, 'data')
OUT_DIR = os.path.join(_HERE, 'docs', '03-analysis', 'data')

NCS_FILE = 'ncs_keywords_in_markdown_results_20260402.xlsx'
TXT_FILE = 'ncs_keywords_in_markdown_results_교과서_results_20260415.xlsx'

# 원본 등급 → 통일 등급 (1 미흡·없음 / 2 형식적 언급 / 3 구체적 대책)
NCS_MAP = {1: 1, 2: 2, 3: 3}
TXT_MAP = {1: 2, 2: 3, 3: 1}

GRADE_LABEL = {1: '미흡·없음', 2: '형식적 언급', 3: '구체적 대책'}

TXT_TOTAL_PAGES = 2055           # 교과서 9권 원본 총 쪽수

# 교과서는 비공개 상용 교재다. 산출 CSV 에는 변환 파이프라인 경로 대신
# 사람이 읽는 제목만 싣는다(대시보드 교재 목록과 동일한 표기).
TXT_TITLES = {
    '반도체기초기술1_크리아트': '반도체 기초기술 1 (크리아트)',
    '반도체기초기술2_크리아트': '반도체 기초기술 2 (크리아트)',
    '반도체기초_렛유인': '반도체 기초 (렛유인)',
    '반도체공정기초_렛유인': '반도체 공정기초 (렛유인)',
    '반도체장비유지보수_충남반도체고': '반도체 장비 유지보수 (충남반도체고)',
    '반도체인프라일반_서울시교육청': '반도체 인프라 일반 (서울시교육청)',
    '반도체_포토에칭_에이치앤지': '반도체 포토에칭 (에이치앤지)',
    '반도체조립검사_에이치앤지': '반도체 조립검사 (에이치앤지)',
    '반도체박막확산_에이치앤지': '반도체 박막확산 (에이치앤지)',
}


def txt_title(fn):
    """교과서 파일 경로를 공개 가능한 제목으로.

    매핑에 없으면 즉시 중단한다. 원본을 그대로 반환하면 비공개 상용 교재의
    변환 파이프라인 경로가 공개 저장소의 CSV 로 새어나간다.
    """
    for key, title in TXT_TITLES.items():
        if fn and key in fn:
            return title
    sys.exit('TXT_TITLES 에 없는 교과서입니다: %r\n'
             '공개 CSV 에 원본 경로가 실리지 않도록 매핑을 추가하십시오.' % fn)


AREAS = ['반도체개발', '반도체제조', '반도체장비', '반도체재료']

YN = {'예', '아니오'}
FN_RE = re.compile(r'^(LM\d|20\d{6}_\d{6}_)')   # NCS: LM…, 교과서: 20260415_143535_…

# 기대값 (회귀 검증용)
#
# truncated_* 는 엑셀 셀 한도(32,767자)에서 잘린 본문의 실측치다. 원본이 남아 있지
# 않아 복구할 수 없으므로 이 값들은 고정 사실이며, 변하면 원본이 교체된 것이다.
# 절단은 NCS 에만 있고 등급3 108쪽 중 12쪽(11.1%)을 차지한다 — 등급1 은 0쪽이라
# 무작위가 아니다. 행으로 세면 1,376건이지만 그건 절단쪽에 키워드가 많이 걸린
# 결과일 뿐이라 페이지로만 센다.
EXPECTED = {
    'ncs': {'rows': 7769, 'pages': 1847, 'books': 86,
            'row_g': {1: 2076, 2: 3465, 3: 2228},
            'page_g': {1: 1270, 2: 469, 3: 108},
            'page_grade_digest': '3461b416055291a5',
            'truncated_pages': 16,
            'truncated_page_g': {1: 0, 2: 4, 3: 12}},
    'txt': {'rows': 981, 'pages': 362, 'books': 9,
            'row_g': {1: 557, 2: 360, 3: 64},
            'page_g': {1: 309, 2: 45, 3: 8},
            'page_grade_digest': '9186a08609cec321',
            'truncated_pages': 0,
            'truncated_page_g': {1: 0, 2: 0, 3: 0}},
}


def as_page(y):
    """페이지 셀을 정수로. openpyxl 이 숫자 셀을 19 또는 19.0 으로 돌려주므로 둘 다 받는다.

    실패 시 None. 다음은 모두 거절한다 — 페이지가 아닌 셀이 조용히 페이지로
    승격되면 고유쪽수가 오염되고 그대로 공개 대시보드에 실린다.
      · nan / inf              (int() 가 예외를 던진다)
      · 과학표기 '1e3', 불리언  (숫자처럼 보이지만 페이지 표기가 아니다)
      · 전각 숫자 '１９'        (원본에 없는 표기 — 통과시키면 오검출 통로가 된다)
    """
    if y is None or isinstance(y, bool):
        return None
    t = str(y).strip()
    if not t or not re.fullmatch(r'[0-9]{1,4}(\.0+)?', t):
        return None
    n = int(float(t))
    return n if 0 < n <= 9999 else None


def parse_row(vals):
    """시트마다 열 구성이 달라, 행 끝의 (사고사례여부, 등급, 등급사유) 3연속을 앵커로 파싱한다.

    앵커는 뒤에서부터 찾는다. 본문·비고 열에 우연히 ('예', '2') 같은 값이 있어도
    실제 꼬리 3연속을 우선하기 위해서다.

    `truncated` 는 본문 열을 찾지 않고 **행의 모든 셀**에서 판정한다. 절단 마커가
    자기 식별적이라 열 위치를 몰라도 되고, 시트마다 열 구성이 다른 이 워크북들에서는
    그게 유일하게 안전한 방법이다 — 교과서 파일은 NCS 와 열 배치가 달라서
    regrade.py 의 고정 열(C/E/F/H)이 맞지 않는다.

    판정은 `v` 가 아니라 **원시 `vals`** 로 한다. `v` 는 strip() 을 거쳐서, 셀 앞뒤에
    공백이 있으면 길이가 32,767 미만으로 줄어 절단을 놓친다.
    """
    v = [None if x is None else str(x).strip() for x in vals]
    truncated = any(is_cell_truncated(y) for y in vals)
    for i in range(len(v) - 1, -1, -1):
        x = v[i]
        if x in YN and i + 1 < len(v) and v[i + 1] in ('1', '2', '3'):
            fn = next((y for y in v[:i] if y and FN_RE.match(y)), None)
            area = next((y for y in v[:i] if y and y.startswith('반도체') and len(y) <= 8), None)
            page = None
            for y in v[:i]:                       # page는 number 뒤에 오므로 마지막 숫자 셀
                n = as_page(y)
                if n is not None:
                    page = n
            return dict(case=x, raw=int(v[i + 1]),
                        reason=v[i + 2] if i + 2 < len(v) else None,
                        fn=fn, area=area, page=page, truncated=truncated)
    return None


def scan(path, gmap, drop_ncs_residue=False):
    wb = load_workbook(path, read_only=True, data_only=True)
    rows, skipped = [], 0
    for sheet in wb.sheetnames:
        for r in wb[sheet].iter_rows(values_only=True):
            if r is None or all(x is None for x in r):
                continue
            if r[0] is not None and str(r[0]).strip() == 'number':
                continue                          # 헤더행
            p = parse_row(r)
            if not p:
                continue
            if drop_ncs_residue and p['fn'] and p['fn'].startswith('LM'):
                skipped += 1                      # 교과서 파일에 남은 NCS 잔여행
                continue
            p['kw'] = sheet
            p['g'] = gmap[p['raw']]
            rows.append(p)
    wb.close()
    if skipped:
        print('  · NCS 잔여행 %d건 제외' % skipped)
    return rows


def page_record(page_rows):
    """한 페이지의 여러 키워드 행을 페이지 단위 레코드 1건으로 축약한다.

    등급은 페이지 속성이지만 채점 버그(`인화` 시트의 총계/내역 불일치)로 12쪽이
    시트마다 다른 등급을 받았다. 스캔 순서에 의존하지 않도록 규칙을 명시한다.

      등급     충돌하면 **가장 낮은 등급**을 택한다(보수적).
               다수결은 쓸 수 없다. 등급이 갈리는 12쪽은 전부 `인화` 등 특정
               시트가 총계를 부풀린 결과인데, 같은 오류 판정이 여러 행에
               중복 기록되거나(예: LM1903060425 p.28 은 4:1) 같은 오류의
               변형이 둘로 갈려(예: p.83 은 "안전 6건"과 "안전 8건") 정확한
               행을 이긴다. 중복 수와 변형 수는 정확성과 무관한 값이다.
               "총계 == 내역 합" 검사도 쓸 수 없다 — 사유 문자열이 상위
               키워드만 싣고 잘리는 정상 행이 41%라 신호가 되지 못한다.
               충돌 시 최저 등급은 이 보고서의 논지(안전 내용 과대평가 방지)와
               같은 방향이고, 판정 근거를 스캔 순서에서 완전히 떼어낸다.
      사고사례  한 행이라도 '예'면 '예' (OR). 이전 구현은 마지막 행만 남겨
               사고사례 8쪽이 CSV에서 전부 '아니오'로 유실됐다
      절단      한 행이라도 잘린 본문을 실었으면 그 페이지의 본문은 잘렸다 (OR).
               등급의 min 과 달리 여기서 보수적인 방향은 OR 이다 — 절단을
               놓치면 오염된 쪽이 깨끗한 층으로 섞여 들어간다.
      등급사유  채택된 등급을 가진 행 중 첫 번째
    """
    grade = min(x['g'] for x in page_rows)
    case = '예' if any(x['case'] == '예' for x in page_rows) else '아니오'
    truncated = any(x.get('truncated') for x in page_rows)
    rep = next(x for x in page_rows if x['g'] == grade)
    return dict(rep, g=grade, case=case, truncated=truncated)


def aggregate(rows, total_pages=None):
    grouped = collections.defaultdict(list)
    for x in rows:
        grouped[(x['fn'], x['page'])].append(x)
    by_page = {k: page_record(v) for k, v in grouped.items()}
    mixed = sum(1 for v in grouped.values() if len(set(x['g'] for x in v)) > 1)
    out = {
        'rows': len(rows),
        'pages': len(by_page),
        'books': len(set(x['fn'] for x in rows)),
        'mixed_grade_pages': mixed,
        'row_g': {g: sum(1 for x in rows if x['g'] == g) for g in (1, 2, 3)},
        'page_g': {g: sum(1 for x in by_page.values() if x['g'] == g) for g in (1, 2, 3)},
        'cases_rows': sum(1 for x in rows if x['case'] == '예'),
        'cases_pages': len(set((x['fn'], x['page']) for x in rows if x['case'] == '예')),
        'cases_books': len(set(x['fn'] for x in rows if x['case'] == '예')),
        # 엑셀 셀 한도(32,767자)에서 잘린 본문. 원본이 남아 있지 않아 복구는 불가이며
        # 이 수치는 층화 분석용이다. 행 단위로 세면 절단쪽에 키워드가 많이 걸려
        # 크게 부풀려지므로(NCS 실측 1,376행 대 16쪽) 페이지 단위로만 센다.
        'truncated_pages': sum(1 for x in by_page.values() if x['truncated']),
        'truncated_page_g': {g: sum(1 for x in by_page.values()
                                    if x['truncated'] and x['g'] == g)
                             for g in (1, 2, 3)},
        # 집계값만 비교하면 페이지 간 등급 재배정이 상쇄되어 검출되지 않는다.
        # (교재,페이지)→등급 전체의 해시를 함께 고정한다.
        'page_grade_digest': hashlib.sha256(
            '\n'.join('%s|%s|%s' % (fn, pg, x['g'])
                       for (fn, pg), x in sorted(by_page.items(),
                                                 key=lambda z: (str(z[0][0]), z[0][1] or 0))
                       ).encode('utf-8')).hexdigest()[:16],
    }
    if total_pages:
        out['total_pages'] = total_pages
        # 검출쪽이 총쪽수를 넘으면(페이지 매핑 오류) 음수 대신 0으로 클램프
        out['undetected_pages'] = max(0, total_pages - out['pages'])
    return out, by_page


def write_atomic(path, write_fn, newline=None, encoding='utf-8'):
    """임시 파일에 쓰고 os.replace 로 교체한다.

    3개 산출물을 순차로 덮어쓰던 중 예외·Ctrl-C 가 나면 새 CSV 와 옛 summary.json
    이 섞인 상태가 남고, 그대로 커밋되면 대시보드 교차검증이 거짓으로 통과한다.
    """
    tmp = path + '.tmp'
    try:
        with open(tmp, 'w', newline=newline, encoding=encoding) as f:
            write_fn(f)
        os.replace(tmp, path)
    except BaseException:
        if os.path.exists(tmp):
            os.remove(tmp)
        raise


def kw_pages(rows):
    """키워드별 고유 검출쪽 수. 대시보드 KW 배열의 pg 컬럼 대조용."""
    d = collections.defaultdict(set)
    for x in rows:
        d[x['kw']].add((x['fn'], x['page']))
    return {k: len(v) for k, v in sorted(d.items())}


def check(name, got, exp):
    bad = []
    for k, v in exp.items():
        if got.get(k) != v:
            bad.append('%s: %s ≠ %s' % (k, got.get(k), v))
    print('  [%s] %s' % (name, '검증 통과' if not bad else '검증 실패 — ' + '; '.join(bad)))
    return not bad


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--out', default=OUT_DIR, help='CSV/JSON 출력 디렉터리')
    ap.add_argument('--data', default=DATA_DIR, help='원본 엑셀 디렉터리')
    ap.add_argument('--force', action='store_true',
                    help='회귀 검증에 실패해도 산출물을 덮어쓴다 (원본이 갱신돼 EXPECTED 를 바꿀 때)')
    args = ap.parse_args()
    os.makedirs(args.out, exist_ok=True)

    print('원본 읽는 중...')
    missing = [f for f in (NCS_FILE, TXT_FILE)
               if not os.path.isfile(os.path.join(args.data, f))]
    if missing:
        sys.exit('원본 엑셀을 찾을 수 없습니다 (%s):\n  %s\n'
                 'data/ 는 .gitignore 대상이라 새로 클론한 환경에는 없습니다. '
                 '원본을 두거나 --data 로 경로를 지정하십시오.'
                 % (args.data, '\n  '.join(missing)))
    ncs = scan(os.path.join(args.data, NCS_FILE), NCS_MAP)
    txt = scan(os.path.join(args.data, TXT_FILE), TXT_MAP, drop_ncs_residue=True)

    ncs_agg, ncs_pages = aggregate(ncs)
    txt_agg, txt_pages = aggregate(txt, TXT_TOTAL_PAGES)
    # 대시보드 KW 배열의 pg(검출쪽) 컬럼이 자기 자신을 근거로 삼지 않도록 독립 산출한다
    ncs_agg['kw_pages'] = kw_pages(ncs)
    txt_agg['kw_pages'] = kw_pages(txt)

    print('\n회귀 검증')
    ok = check('NCS', ncs_agg, EXPECTED['ncs']) & check('교과서', txt_agg, EXPECTED['txt'])

    for name, agg in (('NCS 교재 86권', ncs_agg), ('반도체고 교과서 9권', txt_agg)):
        print('\n%s' % name)
        print('  검출 %d쪽 / %d건 / 등급 혼재 페이지 %d쪽'
              % (agg['pages'], agg['rows'], agg['mixed_grade_pages']))
        for g in (1, 2, 3):
            p, r = agg['page_g'][g], agg['row_g'][g]
            line = '  등급%d %-7s %5d쪽 (%5.1f%%)  행 %5d건 (%5.1f%%)  증폭 %4.1f배' % (
                g, GRADE_LABEL[g], p, p / agg['pages'] * 100 if agg['pages'] else 0,
                r, r / agg['rows'] * 100 if agg['rows'] else 0, r / p if p else 0)
            if 'total_pages' in agg:
                line += '  전체대비 %4.1f%%' % (p / agg['total_pages'] * 100)
            print(line)
        print('  사고사례 판정: %d건 / %d쪽 / %d권'
              % (agg['cases_rows'], agg['cases_pages'], agg['cases_books']))

    # 영역별 (NCS)
    print('\nNCS 영역별 (페이지 기준)')
    print('  %-10s %4s %6s %7s %7s %7s %9s' % ('영역', '권', '검출쪽', '등급1', '등급2', '등급3', '안전관련%'))
    for a in AREAS:
        pl = [x for x in ncs_pages.values() if x['area'] == a]
        g = [sum(1 for x in pl if x['g'] == i) for i in (1, 2, 3)]
        bk = len(set(x['fn'] for x in ncs if x['area'] == a))
        pct = (g[1] + g[2]) / len(pl) * 100 if pl else 0.0
        print('  %-10s %4d %6d %7d %7d %7d %8.1f%%'
              % (a, bk, len(pl), g[0], g[1], g[2], pct))

    # 교재별 (교과서)
    print('\n교과서 교재별 (페이지 기준)')
    for f in sorted(set(x['fn'] for x in txt)):
        pl = [x for x in txt_pages.values() if x['fn'] == f]
        g = [sum(1 for x in pl if x['g'] == i) for i in (1, 2, 3)]
        nm = txt_title(f)          # 콘솔 요약에도 원본 경로를 찍지 않는다
        print('  %-34s 검출 %3d쪽  등급1 %3d  등급2 %3d  등급3 %d'
              % (nm[:34], len(pl), g[0], g[1], g[2]))

    # 등급3 0쪽 교재
    for label, rows_, pages_ in (('NCS', ncs, ncs_pages), ('교과서', txt, txt_pages)):
        bybook = collections.defaultdict(collections.Counter)
        for x in pages_.values():
            bybook[x['fn']][x['g']] += 1
        zero = [b for b, c in bybook.items() if c[3] == 0]
        print('\n%s 등급3(구체적 대책) 0쪽 교재: %d / %d권' % (label, len(zero), len(bybook)))

    if not ok and not args.force:
        sys.exit('\n회귀 검증에 실패해 산출물을 쓰지 않습니다. '
                 '원본이 갱신돼 기대값이 바뀐 것이라면 EXPECTED 를 고치고 다시 실행하거나 '
                 '--force 로 덮어쓰십시오.')

    # 산출물
    # '절단' 은 항상 마지막 열이다. 중간에 끼우면 열 인덱스로 읽는 소비자가 조용히 어긋난다.
    for nm, pages_, cols in (('ncs_pages', ncs_pages, ['영역', '교재', '페이지', '등급', '등급명', '사고사례', '등급사유', '절단']),
                             ('txt_pages', txt_pages, ['교재', '페이지', '등급', '등급명', '사고사례', '등급사유', '절단'])):
        path = os.path.join(args.out, nm + '.csv')

        def _write(f, pages_=pages_, cols=cols):
            w = csv.writer(f)
            w.writerow(cols)
            for (fn, pg), x in sorted(pages_.items(), key=lambda z: (str(z[0][0]), z[0][1] or 0)):
                row = [x['area']] if '영역' in cols else []
                label = fn if '영역' in cols else txt_title(fn)
                row += [label, pg, x['g'], GRADE_LABEL[x['g']], x['case'], x['reason'],
                        '예' if x['truncated'] else '아니오']
                w.writerow(row)

        write_atomic(path, _write, newline='', encoding='utf-8-sig')
        print('저장: %s (%d행)' % (path, len(pages_)))

    summary = {'unified_grades': GRADE_LABEL, 'ncs': ncs_agg, 'textbook': txt_agg}
    sp = os.path.join(args.out, 'summary.json')
    write_atomic(sp, lambda f: json.dump(summary, f, ensure_ascii=False, indent=1))
    print('저장: %s' % sp)

    return 0 if ok else 1


if __name__ == '__main__':
    sys.exit(main())
