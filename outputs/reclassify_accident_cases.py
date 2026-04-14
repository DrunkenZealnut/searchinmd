#!/usr/bin/env python3
"""
사고사례여부 재판정 스크립트

기존 문제:
1. 페이지전체내용 기준 판정 → 같은 페이지의 무관한 행도 전부 "예"
2. 단순 "사고 사례" 단어 매칭 → 교육목표, 학습활동도 "예"

새 기준:
- contents(D열) 기준으로 개별 행 단위 판정
- 구체적 사고사례 = 실제 발생 사건 서술, 사고 통계, 보고된 사례
- 제외: 예방 지침, 과학적 설명, 조사 절차, 교육 목표 등
"""

import openpyxl
import re
import sys
import os
from copy import copy
from datetime import datetime


# ─── 구체적 사고사례 판정 기준 ───

def is_concrete_accident_case(text):
    """
    contents 텍스트에 구체적 사고사례가 포함되어 있는지 판정.

    구체적 사고사례란:
    1. 실제 발생한 사건의 서술 (장소, 시점, 경위, 피해)
    2. 사고 통계 (구체적 수치 포함)
    3. 보고된 사례 참조 (논문, 사례집 인용)

    제외 대상:
    - 예방/방지 지침, 안전 수칙
    - 과학적 성질/특성 설명 (폭발한계, 인화점 등)
    - 조사/분석 절차, 관찰 요령
    - 교육 목표, 학습 활동
    - 건강 유해성 일반 설명 (증상 나열)
    - 표/테이블 데이터
    - 무재해 등 용어 정의
    - "사고가 발생할 수 있다" 같은 가능성/우려 언급
    """
    if not text or len(text.strip()) < 20:
        return False, ''

    text = text.strip()

    # ═══ 1단계: 제외 패턴 (이것에 해당하면 사고사례 아님) ═══
    exclude_patterns = [
        # 교육/학습 관련
        r'학습\s*목표',
        r'수행\s*(내용|순서|tip)',
        r'사고\s*사례를?\s*(교육|학습|조사|수집|분석)',
        r'사고\s*사례에\s*대(한|하여|해)\s*(제한|토론|검색)',
        r'사고\s*사례.*토론',
        r'안전사고\s*사례를?\s*교육',
        r'사고를\s*미연에\s*방지',
        r'안전보건공단.*교재',

        # 예방/방지/대책 (사고사례가 아닌 지침)
        r'사고\s*예방을?\s*위(해|하여)',
        r'안전사고\s*(방지|예방)$',
        r'사고를?\s*방지하기\s*위',

        # 과학적 설명/정의
        r'폭발한계.*Explosion\s*limit',
        r'폭발\(연소\)(상한|하한)계',
        r'연소한계.*flammable\s*limit',
        r'인화점이.*인화성\s*액체',
        r'무재해란',
        r'IUPAC.*명으로는',

        # 조사/관찰 절차
        r'사고\s*현장.*관찰\s*요령',
        r'고지대에서\s*현장\s*전체를\s*관찰',
        r'소손\s*상황을\s*관찰',
        r'연소\s*경로를\s*추측',

        # 표/테이블 (대부분 화학물질 데이터)
        r'^\s*\|.*\|.*\|.*\|.*\|',  # 5개 이상 셀이 있는 테이블 행

        # 이미지 참조
        r'^\[이미지:',

        # 모듈 메타데이터
        r'학습모듈의\s*(개요|목표)',
        r'선수학습',
        r'핵심\s*용어',
        r'재료\s*[·‧]\s*자료',

        # 단순 나열 (항목만)
        r'^(안전|학습)\s*\d\s',
        r'^\*\*(학습|안전|수행)',

        # MSDS 구성 항목 설명
        r'MSDS.*구성\s*항목',
        r'안전성\s*및\s*반응성',
    ]

    for pat in exclude_patterns:
        if re.search(pat, text, re.MULTILINE):
            return False, ''

    # ═══ 2단계: 구체적 사고사례 패턴 매칭 ═══

    evidence = []

    # --- 유형 A: 실제 사건 서술 (가장 확실) ---
    event_patterns = [
        # 구체적 피해 결과 (가장 확실한 지표)
        (r'\d+명이?\s*(사망|부상|중상|경상|입원)', '인명 피해 수치'),
        (r'사망하고\s*\d+명', '사망+부상 수치'),

        # 사고 경위 서술
        (r'(밸\s*브|배관|탱크|설비).*개방.{0,20}유출', '설비 사고 경위'),
        (r'(작업자|근로자)의?\s*(잘못|부주의|실수)로', '사고 원인'),

        # 구체적 지명 + 사고 서술 (지명만으로는 불충분)
        (r'(○○|[가-힣]+)(시|구|동|면|리).{0,20}(산업단지|공장|회사).{0,50}(유출|폭발|사고|화재)', '구체적 지명+사고'),
    ]

    for pat, desc in event_patterns:
        if re.search(pat, text):
            evidence.append(desc)

    # --- 유형 B: 사고 통계 (구체적 수치) ---
    stat_patterns = [
        (r'사고\s*사례집에\s*의하면', '사례집 인용'),
        (r'\d{4}년부터\s*\d{4}년까지.*사고', '기간별 사고 통계'),
        (r'사고\s*중.*\d+건', '사고 건수 통계'),
        (r'사고.*전체\s*사고\s*건수의?\s*약?\s*\d+%', '사고 비율 통계'),
    ]

    for pat, desc in stat_patterns:
        if re.search(pat, text):
            evidence.append(desc)

    # --- 유형 C: 보고된 사례 참조 ---
    report_patterns = [
        (r'피해\s*사례가?\s*보고', '피해 사례 보고'),
        (r'사망사고가?\s*(발생|보고)', '사망사고 보고'),
        (r'사고\s*발\s*생\s*사례가\s*있', '사고 발생 사례 확인'),
        (r'(연구자|논문).*보고한.*사고', '논문 보고 사고'),
        (r'보고된\s*바\s*있다', '사례 보고 확인'),
        (r'급성중독\s*(사망|피해)\s*사례', '급성중독 사례'),
        (r'누출\s*사고로\s*인한\s*급성중독\s*피해\s*사례', '누출 급성중독 사례'),
    ]

    for pat, desc in report_patterns:
        if re.search(pat, text):
            evidence.append(desc)

    # --- 유형 D: 사고 조사/분석 맥락에서의 사건 참조 ---
    investigation_patterns = [
        (r'사고\s*원인.*규명', '사고 원인 규명'),
        (r'(재해|사고)\s*조사\s*결과', '사고 조사 결과'),
        (r'사고\s*경위', '사고 경위'),
    ]

    for pat, desc in investigation_patterns:
        if re.search(pat, text):
            evidence.append(desc)

    # ═══ 3단계: 추가 필터링 (오탐 방지) ═══

    if evidence:
        # "사고가 발생할 수 있다" 같은 가능성 언급은 제외
        possibility_patterns = [
            r'사고가?\s*발생할\s*수\s*있',
            r'사고의?\s*위험이?\s*있',
            r'사고를?\s*일으킬\s*수',
            r'폭발할?\s*우려가?\s*있',
            r'위험에\s*처\s*하거나',
            r'원인이\s*되어서는\s*안\s*된다',
        ]

        # 가능성만 언급하고 실제 사건은 없는 경우
        has_possibility_only = any(re.search(p, text) for p in possibility_patterns)
        strong_evidence = ['인명 피해 수치', '사망+부상 수치', '사례집 인용',
                           '기간별 사고 통계', '사고 건수 통계', '피해 사례 보고',
                           '사망사고 보고', '급성중독 사례', '누출 급성중독 사례',
                           '논문 보고 사고', '사례 보고 확인', '사고 발생 사례 확인',
                           '구체적 지명+사고', '설비 사고 경위', '사고 원인']
        has_real_event = any(e in strong_evidence for e in evidence)

        if has_possibility_only and not has_real_event:
            return False, ''

        # 화학물질 성질 설명에 "폭발" 단어가 있는 경우 제외
        chem_property_patterns = [
            r'인화점.*인화성.*액체',
            r'(가연성|연소).*혼합물',
            r'착화하여\s*연\s*소',
            r'폭발한계',
            r'발화온도',
        ]
        if any(re.search(p, text) for p in chem_property_patterns):
            has_stat = any(e in ['사례집 인용', '기간별 사고 통계', '사고 건수 통계']
                         for e in evidence)
            if not has_stat:
                return False, ''

        # 안전 관리/예방 지침 맥락 (사고라는 단어가 있지만 사고사례가 아닌 것)
        prevention_context = [
            r'사고\s*방\s*지에\s*대한',
            r'안전\s*관리에?\s*규정',
            r'안전하게\s*(저장|공급|관리|처리)',
            r'사고의?\s*위험을?\s*(Zero|제로|최소)',
            r'(폭발성|위험)\s*분위기를?\s*형성하지\s*않도록',
            r'위험성을?\s*평가하기\s*위한',
            r'발생할\s*수\s*있는.*위험성',
            r'고장에\s*의한\s*사고에\s*따른\s*영향을\s*평가',
            r'폐기물.*처리\s*절차\s*준수',
            r'생산\s*계획을?\s*수립',
            r'일정관리.*시스템',
            r'생산집행시스템',
        ]
        if not has_real_event and any(re.search(p, text) for p in prevention_context):
            return False, ''

        # '연도 특정 사고' 근거만 있고, 실제로는 기술 도입/제도 시행인 경우
        if evidence == ['연도 특정 사고'] or (len(evidence) == 1 and evidence[0] == '연도 특정 사고'):
            tech_intro = [
                r'\d{4}년에?\s*(도입|시행|제정|개정|발표)',
                r'\d{4}년.*기술은',
                r'\d{4}년.*시스템',
            ]
            if any(re.search(p, text) for p in tech_intro):
                return False, ''

    if evidence:
        return True, ', '.join(evidence)

    return False, ''


def process_excel(input_path, output_path):
    """Excel 파일의 사고사례여부를 재판정하여 저장"""

    print(f'입력: {input_path}')
    print(f'출력: {output_path}')
    print()

    wb = openpyxl.load_workbook(input_path)

    total_before_yes = 0
    total_after_yes = 0
    total_changed_to_no = 0
    total_changed_to_yes = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [str(c.value) for c in ws[1]]

        # 사고사례여부 컬럼 찾기
        sago_col = None
        reason_col = None
        for i, h in enumerate(headers):
            if '사고사례' in str(h):
                sago_col = i + 1
            if '등급사유' in str(h):
                reason_col = i + 1

        if not sago_col:
            continue

        # contents는 항상 4번째 열
        contents_col = 4

        sheet_before = 0
        sheet_after = 0
        sheet_to_no = 0
        sheet_to_yes = 0

        for row_idx in range(2, ws.max_row + 1):
            old_val = ws.cell(row=row_idx, column=sago_col).value
            contents = str(ws.cell(row=row_idx, column=contents_col).value or '')

            if old_val == '예':
                total_before_yes += 1
                sheet_before += 1

            # 재판정
            is_case, evidence = is_concrete_accident_case(contents)
            new_val = '예' if is_case else '아니오'

            # 변경 사항 기록
            if old_val == '예' and new_val == '아니오':
                total_changed_to_no += 1
                sheet_to_no += 1
            elif old_val == '아니오' and new_val == '예':
                total_changed_to_yes += 1
                sheet_to_yes += 1

            if new_val == '예':
                total_after_yes += 1
                sheet_after += 1

            # 셀 업데이트
            ws.cell(row=row_idx, column=sago_col).value = new_val

            # 등급사유에 사고사례 근거 업데이트
            if reason_col:
                old_reason = str(ws.cell(row=row_idx, column=reason_col).value or '')
                # 기존 사고사례: 부분 제거
                if '| 사고사례:' in old_reason:
                    base_reason = old_reason.split('| 사고사례:')[0].strip()
                elif '사고사례:' in old_reason:
                    parts = old_reason.split('사고사례:')
                    base_reason = parts[0].strip().rstrip('|').strip()
                else:
                    base_reason = old_reason

                if is_case and evidence:
                    new_reason = f'{base_reason} | 사고사례: {evidence}'
                else:
                    new_reason = base_reason

                ws.cell(row=row_idx, column=reason_col).value = new_reason

        if sheet_before > 0 or sheet_after > 0:
            print(f'  [{sheet_name:>10}] 기존={sheet_before:>3} → 변경후={sheet_after:>3}  '
                  f'(예→아니오: {sheet_to_no}, 아니오→예: {sheet_to_yes})')

    print()
    print(f'===== 전체 요약 =====')
    print(f'기존 사고사례=예: {total_before_yes}건')
    print(f'변경 후 사고사례=예: {total_after_yes}건')
    print(f'예→아니오 (오탐 제거): {total_changed_to_no}건')
    print(f'아니오→예 (신규 발견): {total_changed_to_yes}건')
    print(f'정밀도 향상: {total_before_yes}건 → {total_after_yes}건 '
          f'({(total_before_yes - total_after_yes) / total_before_yes * 100:.1f}% 감소)')

    # 저장
    wb.save(output_path)
    print(f'\n저장 완료: {output_path}')


if __name__ == '__main__':
    input_file = '/Users/zealnutkim/Documents/청년노동자인권센터/2026년/교과서 기초연구/ncs_keywords_in_markdown_results_20260402.xlsx'

    # 출력 파일명: 원본_재판정_날짜.xlsx
    base_dir = os.path.dirname(input_file)
    base_name = os.path.splitext(os.path.basename(input_file))[0]
    timestamp = datetime.now().strftime('%Y%m%d')
    output_file = os.path.join(base_dir, f'{base_name}_재판정_{timestamp}.xlsx')

    process_excel(input_file, output_file)
