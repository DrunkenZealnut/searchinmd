#!/usr/bin/env python3
"""
Excel 첫 번째 탭의 각 행에서 D열(페이지 번호)를 읽고,
해당 마크다운 파일에서 그 페이지의 전체 내용을 E열에 추가합니다.
"""

import os
import sys
import json
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from page_utils import nfc, find_md_and_meta, build_page_map, extract_page_content

# === 설정 ===
EXCEL_PATH = "/Users/zealnutkim/Documents/청년노동자인권센터/2026년/교과서 기초연구/ncs_keywords&reworkinglist.xlsx"
NCS_BASES = [
    "/Users/zealnutkim/Documents/청년노동자인권센터/교과서/교과서/NCS",
    "/Users/zealnutkim/Documents/DEV/SafeFactory/documents/semiconductor/ncs/data",
]
OUTPUT_PATH = EXCEL_PATH.replace(".xlsx", "_with_fullpage.xlsx")
EXCEL_MAX_CHARS = 32767


def main():
    print(f"Excel 파일 읽는 중: {EXCEL_PATH}")
    wb = load_workbook(EXCEL_PATH)
    ws = wb[wb.sheetnames[0]]
    print(f"시트: '{wb.sheetnames[0]}', 행: {ws.max_row}, 열: {ws.max_column}")

    if ws.cell(row=2, column=5).value is None:
        ws.cell(row=2, column=5, value='페이지전체내용')

    file_cache = {}
    processed = 0
    skipped = 0
    errors = []

    for row_idx in range(3, ws.max_row + 1):
        filename = ws.cell(row=row_idx, column=2).value
        page_val = ws.cell(row=row_idx, column=4).value

        if not filename or page_val is None:
            skipped += 1
            continue

        try:
            target_page = int(page_val)
        except (ValueError, TypeError):
            skipped += 1
            continue

        filename = nfc(str(filename))

        if filename not in file_cache:
            md_path, meta_path = find_md_and_meta(NCS_BASES, filename)
            if md_path and os.path.exists(md_path):
                with open(md_path, encoding='utf-8') as f:
                    content = nfc(f.read())
                lines = content.split('\n')

                metadata = None
                if meta_path and os.path.exists(meta_path):
                    with open(meta_path, encoding='utf-8') as f:
                        metadata = json.load(f)

                page_map = build_page_map(lines, metadata)
                file_cache[filename] = (lines, page_map)
            else:
                file_cache[filename] = None
                if md_path is None:
                    errors.append(f"Row {row_idx}: 파일 못찾음 - {filename}")

        cached = file_cache.get(filename)
        if cached is None:
            skipped += 1
            continue

        lines, page_map = cached
        full_content = extract_page_content(lines, page_map, target_page, 1)

        if len(full_content) > EXCEL_MAX_CHARS:
            full_content = full_content[:EXCEL_MAX_CHARS - 3] + '...'

        ws.cell(row=row_idx, column=5, value=full_content)
        processed += 1

        if processed % 500 == 0:
            print(f"  처리 중... {processed}행 완료")

    ws.column_dimensions['E'].width = 100
    wb.save(OUTPUT_PATH)
    print(f"\n=== 완료 ===")
    print(f"처리 행: {processed}")
    print(f"건너뛴 행: {skipped}")
    print(f"파일 캐시: {len(file_cache)}개 파일")
    if errors:
        print(f"오류: {len(errors)}건")
        for e in errors[:10]:
            print(f"  {e}")
    print(f"저장: {OUTPUT_PATH}")


if __name__ == '__main__':
    main()
