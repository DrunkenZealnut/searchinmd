#!/usr/bin/env python3
"""
resegment.py — NCS 워크북의 '페이지' 라벨(목차 블록)을 원본 PDF 의 실제 쪽으로 다시 배치하고
페이지 단위 등급을 재집계한다.

    python3.13 resegment.py --pdf-root /path/to/ncs/pdfs            # 기본: 환경변수 NCS_PDF_ROOT
    python3.13 resegment.py --pdf-root ... --limit 3                # 앞 3권만 (디버그)

왜 필요한가. 2026-04 검색 당시 마크다운의 페이지 마커는 `_meta.json` 목차에서 유도한 것이라
목차 항목 하나가 여러 쪽을 한 '페이지' 라벨로 묶었다(32,767자 라벨 16개, 8,000자 초과 38개,
『반도체 장비 안전관리』는 945건이 20개 라벨). 외부감사(2026-09-04) C1. 현재 저장소의 마크다운
마커도 89권 중 23권만 쪽 단위라 재검색으로는 풀리지 않는다.

무엇을 하나. (1) PDF 쪽 텍스트(PyMuPDF)와 마크다운 줄을 문자 3-gram 포함률 + 단조 DP 로
정렬해 줄→실제 쪽 대응을 만들고(쪽 단위 마커가 촘촘한 교재 — 밀도 ≥ DENSE_MARKER_RATIO — 는 마커를 쓰되
마커가 빠진 쪽은 hybrid_pages 가 DP 배정으로 쪼갠다, Act-3), (2) 워크북 검출 행의 매칭 문장을 마크다운 줄에 붙여 실제
쪽으로 옮기고, (3) 실제 쪽 본문에 현행 등급 규칙(`regrade.grade_page` 재현 기준선)을 다시
적용해 (교재, 쪽) 단위로 집계한다. 검출 행 자체는 그대로다 — 쪽 배치와 등급만 바뀐다.

하지 않는 것. 검색 재실행, 새 적중, 등급 규칙 변경, 대시보드 갱신. PDF·마크다운이 없는
교재는 구 라벨·구 등급을 그대로 두고 `unresolved` 로 센다.

산출물 (docs/03-analysis/data/): ncs_pages_reseg.csv, reseg_summary.json.
줄→쪽 대응은 data/markdown/ncs_paged/<LM코드>.pages.json 과 행 대응표 rows_map.csv (둘 다 gitignore) 에 남긴다.
정렬 정확도는 쪽 단위 마커를 이미 가진 교재에서 마커와 대조해 summary 의 alignment_check 에
기록한다 — 이 값이 없으면 산출물의 정밀도를 주장할 수 없다.
"""
import argparse
import collections
import csv
import datetime
import glob
import hashlib
import json
import os
import re
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
import regrade as RG  # noqa: E402
from recount_grades import as_page, write_atomic  # noqa: E402
from page_utils import nfc, GRADE_LABEL, PAGE_MARKER_RE as MARKER_RE  # noqa: E402

_STRIP = re.compile(r'[\s#*_>|\[\]()!`\-–·•○●◆◇■□▶▷※,.:;]+')
CODE_RE = re.compile(r'(LM\d{10})')
DEFAULT_MD_ROOT = os.path.join(HERE, 'data', 'markdown', 'ncs')
DEFAULT_WORKBOOK = os.path.join(HERE, 'data', RG.NCS_FILE)
DEFAULT_OUT = os.path.join(HERE, 'docs', '03-analysis', 'data')
DEFAULT_PAGED = os.path.join(HERE, 'data', 'markdown', 'ncs_paged')
DENSE_MARKER_RATIO = 0.8        # 마커 수가 PDF 쪽수의 80% 이상이면 쪽 단위 마커로 보고 정렬을 검증한다
COL_NUMBER, COL_AREA, COL_CONTENTS, COL_CASE, COL_REASON = 0, 1, 3, 6, 8          # 나머지 열 위치는 regrade.py 와 공유
COL_FILENAME, COL_PAGE, COL_GRADE = RG.COL_FILENAME, RG.COL_PAGE, RG.COL_GRADE
MIN_KEY_CHARS = 10              # 행→줄 매칭 키의 최소 길이 — 이보다 짧은 마지막 줄은 전체 본문 키로 대신한다
BASELINE_KW = dict(word_boundary=False, normalize=False)   # 현행 규칙 재현 기준선 (D1·D2 끄기) — regrade_page 와 meta.rule 이 같이 쓴다

# 회귀 가드 — 보고서가 인용한 수치. 재실행이 여기서 어긋나면 --force 없이는 쓰지 않는다 (regrade.py 의 EXPECTED 와 같은 역할).
# digest 는 (교재, 쪽, 등급) 전체의 지문이라 총계가 같아도 쪽→등급 재배정을 잡는다.
EXPECTED = {'pages': 2189, 'page_g': {'1': 1519, '2': 525, '3': 145}, 'books': 86, 'unresolved_pages': 51,
            'digest': '20855b3bc05d906b',                        # 2026-09-06 Act-3(마커 결손 하이브리드, 앵커 보정) 실행의 (교재, 쪽, 등급) 지문
            'kw_pages_digest': '4eae1c027aa25261', 'case_pages_digest': '7c97d3ceb40a64ff',   # 키워드별 검출 쪽·사고사례 쪽의 정체 (등급 지문이 못 보는 것)
            'cases_pages': 13, 'cases_books': 5, 'moved_rows': 4270, 'unmatched_rows': 94,   # 보고서·TODOS 가 인용하는 나머지 수치 —
            'label_fallback_pages': 24, 'hybrid_lines': 1575,   # match_rows·사고사례 OR·마커 결손 보정이 바뀌면 지문은 그대로여도 여기서 잡힌다
            'hybrid_emptied_marker_pages': 0,                    # 보정이 마커 쪽을 지우면 0 이 아니게 된다
            'alignment_overall': {'lines': 21711, 'exact': 18142, 'near': 20613, 'all_lines': 32486, 'all_exact': 25219, 'all_near': 29329,
                                  'nogap_lines': 19698, 'nogap_exact': 17456, 'nogap_near': 19397},
            'match_stats': {'overflow': 293, 'ambiguous': 1118, 'partial': 118}}   # 문서가 인용하는 자기 검증·약한 배정 수치 — 정의가 바뀌면 여기서 잡힌다


# ---------------------------------------------------------------- 텍스트 정규화·정렬
def norm_text(s):
    """NFC 로 통일하고 공백·마크다운 기호·문장부호를 지운다. 비교 전용."""
    return _STRIP.sub('', nfc(s or ''))


def grams(s, n=3):
    return {s[i:i + n] for i in range(len(s) - n + 1)}


def align_lines(lines, pages_text, min_len=12, min_contain=0.5, window=6, jump_pen=0.06, far_pen=1.0):
    """마크다운 줄 → PDF 쪽(1-based) 단조 정렬.

    후보 줄: 정규화 길이 >= min_len 이고 어느 쪽에든 3-gram 포함률 >= min_contain. 마커 줄 제외.
    DP: 같은 쪽 유지 0, 앞으로 d쪽(<= window) 점프는 jump_pen*d, 더 먼 앞 점프는 far_pen,
    뒤로는 불가. 점수 = 포함률 합 - 벌점 합. 반환: {줄 idx: 쪽}.
    """
    pg = [grams(norm_text(t)) for t in pages_text]
    n_pages = len(pg)
    if n_pages == 0:
        return {}
    elig = []
    for i, line in enumerate(lines):
        if MARKER_RE.search(line):
            continue
        k = norm_text(line)
        if len(k) < min_len:
            continue
        g = grams(k)
        if not g:
            continue
        c = [len(g & pg[p]) / len(g) for p in range(n_pages)]
        if max(c) < min_contain:
            continue
        elig.append((i, c))
    if not elig:
        return {}
    neg = float('-inf')
    prev, back = None, []
    for i, c in elig:
        if prev is None:
            cur = [c[p] - 1e-3 * p for p in range(n_pages)]     # 동점이면 앞 쪽
            bp = list(range(n_pages))
        else:
            cur, bp = [neg] * n_pages, [0] * n_pages
            best_far, best_far_q = neg, 0
            for p in range(n_pages):
                q_far = p - window - 1
                if q_far >= 0 and prev[q_far] > best_far:
                    best_far, best_far_q = prev[q_far], q_far
                cand, cq = prev[p], p
                for d in range(1, window + 1):
                    q = p - d
                    if q < 0:
                        break
                    v = prev[q] - jump_pen * d
                    if v > cand:
                        cand, cq = v, q
                if best_far - far_pen > cand:
                    cand, cq = best_far - far_pen, best_far_q
                cur[p], bp[p] = cand + c[p], cq
        back.append(bp)
        prev = cur
    assigned = {}
    p = max(range(n_pages), key=lambda x: prev[x])
    for j in range(len(elig) - 1, -1, -1):
        assigned[elig[j][0]] = p + 1
        p = back[j][p]
    return assigned


def propagate(n, assigned):
    """미정렬 줄은 직전 정렬 줄의 쪽을, 문서 첫머리는 다음 정렬 줄의 쪽을 물려받는다."""
    out, cur = [None] * n, None
    for i in range(n):
        if i in assigned:
            cur = assigned[i]
        out[i] = cur
    nxt = None
    for i in range(n - 1, -1, -1):
        if out[i] is not None:
            nxt = out[i]
        elif nxt is not None:
            out[i] = nxt
    return out


def page_texts(lines, line_pages):
    """쪽 → 본문(그 쪽에 속한 줄, 마커 줄 제외)."""
    buckets = collections.defaultdict(list)
    for line, page in zip(lines, line_pages):
        if page is None or MARKER_RE.search(line):
            continue
        buckets[page].append(line)
    return {p: '\n'.join(v) for p, v in buckets.items()}


def marker_positions(lines):
    """[(줄 idx, 마커 쪽)] — 마커 줄 목록. marker_pages·hybrid_pages 가 같은 스캔을 쓴다."""
    return [(i, int(m.group(1))) for i, line in enumerate(lines) if (m := MARKER_RE.search(line))]


def marker_pages(lines):
    """마커로 줄→쪽. 마커 줄을 정렬 결과처럼 취급해 propagate 로 채운다 (첫 마커 앞의 줄은 첫 마커의 쪽)."""
    return propagate(len(lines), dict(marker_positions(lines)))


def hybrid_pages(lines, marker_lp, dp_lp, n_pages):
    """마커 우선 교재의 마커 결손 보정 (Act-3, 연구 책임자 결정 2026-09-06).

    변환기가 마커를 빠뜨린 쪽의 줄은 marker_pages 가 앞 마커의 쪽에 붙인다. 마커 N 다음 마커가 N+2 이상이면
    (사이에 마커 없는 쪽이 있으면) 그 사이 줄은 DP 배정이 [N, 다음 마커-1] 안에 들 때 DP 쪽을 쓰고, DP 근거가
    없는 줄은 바로 앞 줄의 쪽을 물려받는다(구간 안 단조). 마지막 마커 뒤는 PDF 끝쪽까지 같은 규칙.
    마커가 빠지지 않은 구간(다음 마커 = N+1)과 첫 마커 앞 줄은 그대로 — 변환기 마커가 DP 보다 정확하다.

    앵커: 구간 첫 본문 줄의 DP 가 마커 쪽 N 보다 앞서 있으면(DP 드리프트 — 마커 앞 줄에서 전파된 값) 그 차이를
    구간 전체의 DP 에서 빼고 판정한다. 그러지 않으면 마커 N 직후 줄이 N+k 로 가 마커가 찍힌 실제 쪽 N 이 본문
    0줄로 비는데(출하 전 리뷰 F1: 실측 17쪽), 마커는 그 줄이 N 에 있다는 직접 증거다. DP 의 절대값이 아니라
    증가분만 쓰는 셈이다.
    반환: 줄→쪽 목록 (marker_lp 와 같은 길이).
    """
    out = list(marker_lp)
    marks = marker_positions(lines)
    if not marks:
        return out
    segs = [(a + 1, b, pa, pb - 1) for (a, pa), (b, pb) in zip(marks, marks[1:])]
    segs.append((marks[-1][0] + 1, len(lines), marks[-1][1], n_pages))
    for start, end, lo, hi in segs:
        if hi - lo < 1:                                   # 빠진 쪽이 없는 구간
            continue
        first = next((dp_lp[j] for j in range(start, min(end, len(dp_lp))) if dp_lp[j] is not None and norm_text(lines[j])), None)
        shift = max(0, first - lo) if first is not None else 0    # 구간 첫 본문 줄의 DP 드리프트
        cur = lo
        for j in range(start, end):
            p = dp_lp[j] if j < len(dp_lp) else None
            if p is not None:
                p -= shift
                if cur <= p <= hi:
                    cur = p
            out[j] = cur
    return out


def emptied_marker_pages(lines, line_pages):
    """마커 쪽 N 아래에 본문 줄이 있는데 최종 줄→쪽에 N 이 하나도 없는 쪽 수 — 결손 보정(hybrid_pages)이 마커 쪽을 지웠는지 드러낸다."""
    marks = marker_positions(lines)
    n = 0
    for (a, pa), (b, _) in zip(marks, marks[1:] + [(len(lines), None)]):
        body = [j for j in range(a + 1, b) if norm_text(lines[j])]
        if body and all(line_pages[j] != pa for j in body):
            n += 1
    return n


def gap_lines(lines, n_pages):
    """마커 결손 구간(hybrid_pages 가 손대는 줄)의 idx 집합 — 자기 검증에서 '정답이 마커 전파값이라 정의상 틀린' 줄을 빼기 위해."""
    marks = marker_positions(lines)
    if not marks:
        return set()
    segs = [(a + 1, b, pa, pb - 1) for (a, pa), (b, pb) in zip(marks, marks[1:])]
    segs.append((marks[-1][0] + 1, len(lines), marks[-1][1], n_pages))
    return {j for start, end, lo, hi in segs if hi - lo >= 1 for j in range(start, end)}


def check_alignment(lines, assigned, exclude=()):
    """마크다운의 마커를 정답으로 정렬 결과를 대조한다. {lines, exact, near(±1)}. exclude 의 줄 idx 는 분모에서 뺀다."""
    truth, cur = {}, None
    for i, line in enumerate(lines):
        m = MARKER_RE.search(line)
        if m:
            cur = int(m.group(1))
            continue
        truth[i] = cur
    n = exact = near = 0
    for i, p in assigned.items():
        t = truth.get(i)
        if t is None or i in exclude:
            continue
        n += 1
        exact += (p == t)
        near += (abs(p - t) <= 1)
    return {'lines': n, 'exact': exact, 'near': near}


def check_alignment_all(lines, line_pages):
    """propagate 까지 마친 줄→쪽을 마커와 대조한다 — 후보 줄만 세는 check_alignment 와 달리 본문이 있는 모든 줄이 분모다.
    {all_lines, all_exact, all_near}. 첫 마커 앞 줄·마커 줄·빈 줄은 뺀다."""
    truth, cur = {}, None
    for i, line in enumerate(lines):
        m = MARKER_RE.search(line)
        if m:
            cur = int(m.group(1))
            continue
        if norm_text(line):
            truth[i] = cur
    n = exact = near = 0
    for i, t in truth.items():
        p = line_pages[i] if i < len(line_pages) else None
        if t is None or p is None:
            continue
        n += 1
        exact += (p == t)
        near += (abs(p - t) <= 1)
    return {'all_lines': n, 'all_exact': exact, 'all_near': near}


# ---------------------------------------------------------------- 워크북 행 → 줄
def sentence_key(contents):
    """contents 의 마지막 줄(제목 맥락 제거)을 키로. MIN_KEY_CHARS 미만이면 호출부가 전체 키로 대신한다. 반환 (짧은 키, 전체 키)."""
    body = (contents or '').strip().split('\n')
    sent = body[-1] if len(body) > 1 else body[0]
    return norm_text(sent), norm_text(contents)


def match_rows(rows, lines, stats=None):
    """행별 줄 idx (없으면 None). 같은 (시트, 키) 의 행은 문서 순서로 서로 다른 적중에 배정.

    stats 에 dict 를 주면 배정의 근거가 약한 행을 센다 — overflow(적중보다 행이 많아 마지막 적중을 재사용),
    ambiguous(적중이 둘 이상이라 문서 순서 가정에 기댐), partial(줄 전체가 아니라 긴 줄의 일부로 적중).
    """
    nl = [norm_text(x) for x in lines]
    st = {'overflow': 0, 'ambiguous': 0, 'partial': 0}

    def hits_for(key):
        for size in (len(key), 80, 50, 30):
            k = key[:size]
            if len(k) < MIN_KEY_CHARS:
                break
            h = [i for i, x in enumerate(nl) if k in x]
            if h:
                return h
        return []

    cache, ptr, out = {}, collections.Counter(), []
    for r in rows:
        short, full = sentence_key(r.get('contents'))
        key = short if len(short) >= MIN_KEY_CHARS else full
        if key not in cache:
            if len(short) >= MIN_KEY_CHARS:
                cache[key] = hits_for(short)
            else:                                   # '안전 · 유의 사항' 같은 짧은 정형구: 줄 전체가 같을 때만
                h = [i for i, x in enumerate(nl) if x == short and x]
                cache[key] = h or hits_for(full)
        h = cache[key]
        if not h:
            out.append(None)
            continue
        gk = (r.get('sheet'), key)          # 짧은 정형구는 전체 키(제목 맥락 포함)로 묶여 맥락이 다르면 각각 첫 등장 줄부터 센다.
                                            # (시트, 정형구) 로 묶는 대안은 Act-2 기준 실측 2,173→2,178쪽(등급1 +5, 등급3 불변, 3권) — 연구 책임자 결정(2026-09-06): 현행 유지 (TODOS.md P2)
        n = ptr[gk]
        ptr[gk] += 1
        i = h[n] if n < len(h) else h[-1]
        st['overflow'] += n >= len(h)
        st['ambiguous'] += len(h) > 1
        st['partial'] += nl[i] != short
        out.append(i)
    if stats is not None:
        stats.update(st)
    return out


def regrade_page(text):
    """현행 규칙 재현 기준선 (D1·D2 끄고). 반환 (등급, 안전수, 조치수, 사유)."""
    return RG.grade_page(text, **BASELINE_KW)


# ---------------------------------------------------------------- 집계
def page_grade_digest(books):
    """(교재, 쪽, 등급) 전체의 지문. 순서와 무관하고 한 쪽의 등급만 바뀌어도 달라진다."""
    lines = sorted('%s\t%s\t%s' % (name, pg, rec.get('grade')) for name, b in books.items() for pg, rec in b['pages'].items())
    return hashlib.sha256('\n'.join(lines).encode('utf-8')).hexdigest()[:16]


def check_expected(summary, expected=None):
    """EXPECTED 와의 불일치 목록. 비어 있으면 통과. digest 가 None 이면(아직 미고정) 지문은 비교하지 않는다."""
    e = EXPECTED if expected is None else expected
    bad = []
    for k in ('pages', 'books'):
        if summary.get(k) != e.get(k):
            bad.append('%s: %s != %s' % (k, summary.get(k), e.get(k)))
    if summary.get('page_g') != e.get('page_g'):
        bad.append('page_g: %s != %s' % (summary.get('page_g'), e.get('page_g')))
    if (summary.get('unresolved') or {}).get('pages') != e.get('unresolved_pages'):
        bad.append('unresolved_pages: %s != %s' % ((summary.get('unresolved') or {}).get('pages'), e.get('unresolved_pages')))
    if e.get('digest') and summary.get('page_grade_digest') != e['digest']:
        bad.append('digest: %s != %s' % (summary.get('page_grade_digest'), e['digest']))
    for k in ('kw_pages_digest', 'case_pages_digest'):    # 등급 지문이 못 보는 것 — 키워드별 검출 쪽·사고사례 쪽의 정체
        if e.get(k) and summary.get(k) != e[k]:
            bad.append('%s: %s != %s' % (k, summary.get(k), e[k]))
    special = ('pages', 'books', 'page_g', 'unresolved_pages', 'digest', 'alignment_overall', 'match_stats', 'kw_pages_digest', 'case_pages_digest')
    for k in e:                                          # 나머지 키는 전부 그대로 대조 — EXPECTED 에 더한 키가 가드 밖에 남을 통로가 없다
        if k not in special and summary.get(k) != e[k]:
            bad.append('%s: %s != %s' % (k, summary.get(k), e[k]))
    if 'alignment_overall' in e and (summary.get('alignment_check') or {}).get('overall') != e['alignment_overall']:
        bad.append('alignment_overall: %s != %s' % ((summary.get('alignment_check') or {}).get('overall'), e['alignment_overall']))
    if 'match_stats' in e and summary.get('match_stats') != e['match_stats']:
        bad.append('match_stats: %s != %s' % (summary.get('match_stats'), e['match_stats']))
    return bad


def aggregate(books):
    """books: {교재: {area, status, rows, moved_rows, unmatched_rows, old_labels, align, pages: {쪽: {...}}}}"""
    page_g, areas = collections.Counter(), {}
    kw_pages, cases_pages, cases_books = collections.Counter(), 0, 0
    unresolved = {'books': 0, 'pages': 0, 'rows': 0}
    align_books, align_tot = {}, {'lines': 0, 'exact': 0, 'near': 0, 'all_lines': 0, 'all_exact': 0, 'all_near': 0, 'nogap_lines': 0, 'nogap_exact': 0, 'nogap_near': 0}
    n_pages, label_fallback, match_tot, fb_on_text, hybrid_tot, emptied_tot = 0, 0, collections.Counter(), 0, 0, 0
    for name, b in books.items():
        hybrid_tot += b.get('hybrid_lines', 0)
        emptied_tot += b.get('emptied_marker_pages', 0)
        a = areas.setdefault(b['area'], {'books': 0, 'pages': 0, 'page_g': collections.Counter()})
        a['books'] += 1
        had_case = False
        for pg, rec in b['pages'].items():
            n_pages += 1
            page_g[str(rec['grade'])] += 1
            a['pages'] += 1
            a['page_g'][str(rec['grade'])] += 1
            for kw in rec.get('kws', ()):
                kw_pages[kw] += 1
            if rec.get('case'):
                cases_pages += 1
                had_case = True
            if rec.get('source') in ('label', 'text-fallback') and b.get('status') != 'unresolved':
                label_fallback += 1                        # 해결 교재인데 매칭 행이 하나도 없어 구 라벨로만 존재하는 쪽
            if rec.get('matched'):
                fb_on_text += rec.get('fallback_rows', 0)  # 매칭 행이 있는 쪽에 구 라벨 번호로 합류한 미매칭 행 — kws·case 에는 안 들어갔다
        cases_books += had_case
        if b.get('status') == 'unresolved':
            unresolved['books'] += 1
            unresolved['pages'] += len(b['pages'])
            unresolved['rows'] += b.get('rows', 0)
        if b.get('align'):
            align_books[name] = b['align']
            for k in align_tot:
                align_tot[k] += b['align'].get(k, 0)
        match_tot.update(b.get('match') or {})
    for a in areas.values():
        a['page_g'] = {g: a['page_g'].get(g, 0) for g in ('1', '2', '3')}
    return {
        'pages': n_pages, 'books': len(books),
        'page_g': {g: page_g.get(g, 0) for g in ('1', '2', '3')},
        'cases_pages': cases_pages, 'cases_books': cases_books,
        'areas': areas, 'unresolved': unresolved,
        'moved_rows': sum(b.get('moved_rows', 0) for b in books.values()),
        'unmatched_rows': sum(b.get('unmatched_rows', 0) for b in books.values()),
        'label_fallback_pages': label_fallback,       # 해결 교재 안에서 구 라벨로만 존재하는 쪽 (출처 label + text-fallback)
        'match_stats': {k: match_tot.get(k, 0) for k in ('overflow', 'ambiguous', 'partial')},   # match_rows 의 약한 배정 (해결 교재 합)
        'hybrid_lines': hybrid_tot,                        # 마커 교재에서 마커 결손 보정으로 쪽이 바뀐 본문 줄 수 (교재별 hybrid_lines 합)
        'hybrid_emptied_marker_pages': emptied_tot,        # 마커가 찍힌 쪽인데 본문 줄이 하나도 남지 않은 쪽 — 0 이어야 정상
        'fallback_rows_on_text_pages': fb_on_text,
        'page_grade_digest': page_grade_digest(books),
        'kw_pages_digest': hashlib.sha256('\n'.join('%s\t%d' % kv for kv in sorted(kw_pages.items())).encode('utf-8')).hexdigest()[:16],
        'case_pages_digest': hashlib.sha256('\n'.join(sorted('%s\t%s' % (name, pg) for name, b in books.items() for pg, rec in b['pages'].items() if rec.get('case'))).encode('utf-8')).hexdigest()[:16],
        'kw_pages': dict(kw_pages),
        'alignment_check': {'books': len(align_books), 'overall': align_tot, 'per_book': align_books},
    }


# ---------------------------------------------------------------- 입력
def load_rows(path, loader=None):
    """워크북 전체 행. 열은 위치로 읽는다 (regrade.py 와 같은 이유 — 시트마다 헤더가 다르다).
    loader 는 시험용 주입점 (기본 openpyxl.load_workbook)."""
    if loader is None:
        from openpyxl import load_workbook as loader
    wb = loader(path, read_only=True, data_only=True)
    rows = []
    for sn in wb.sheetnames:
        for row in wb[sn].iter_rows(values_only=True):
            if not row or len(row) < RG.N_COLS or row[COL_FILENAME] is None:
                continue
            if str(row[COL_NUMBER]).strip() == 'number' or str(row[COL_FILENAME]) == 'filename':
                continue
            label, grade = as_page(row[COL_PAGE]), as_page(row[COL_GRADE])    # nan/inf/불리언/과학표기는 None (recount 규칙)
            rows.append({'sheet': sn, 'area': str(row[COL_AREA] or ''),
                         'filename': nfc(str(row[COL_FILENAME])),
                         'contents': str(row[COL_CONTENTS] or ''), 'label': label,
                         'case': str(row[COL_CASE] or '').strip() == '예', 'grade': grade,
                         'reason': str(row[COL_REASON] or '')})
    wb.close()
    return rows


def index_files(root, pattern):
    out = collections.defaultdict(list)
    for f in sorted(glob.glob(os.path.join(root, '**', pattern), recursive=True)):     # glob 순서는 파일시스템 의존 — 정렬해 결정적으로

        m = CODE_RE.search(nfc(os.path.basename(f)))
        if m:
            out[m.group(1)].append(f)
    return out


def pick_md(code, filename, md_index):
    """같은 코드의 마크다운이 여럿이면 워크북 파일명과 가장 긴 공통 접두를 가진 것, 동점이면 마커가 많은 것.

    동점은 실제로 있다 — 'MI 장비 운영.md'(마커 0)와 'MI_장비_운영.md'(마커 84)처럼 공백/밑줄만 다른 중복.
    glob 순서에 맡기면 마커 없는 쪽을 고를 수 있어(Gap G16) 마커 수를 2차 기준으로 쓴다.
    """
    cands = md_index.get(code) or []
    if not cands:
        return None
    if len(cands) == 1:
        return cands[0]

    def prefix(f):
        base = nfc(os.path.basename(f))
        base = re.sub(r'^\d{8}_\d{6}_', '', base)
        n = 0
        for a, b in zip(base.replace(' ', '_'), filename):
            if a != b:
                break
            n += 1
        return n

    def markers(f):
        try:
            with open(f, encoding='utf-8') as fh:
                return sum(1 for line in fh if MARKER_RE.search(line))
        except OSError:
            return 0
    return max(cands, key=lambda f: (prefix(f), markers(f)))


def unresolved_pages(rows):
    """마크다운·PDF 가 없는 교재: 구 라벨을 쪽으로, 등급은 행 최저(recount 규칙), 사유는 첫 행."""
    pages = {}
    for r in rows:
        if r['label'] is None:
            continue
        rec = pages.setdefault(r['label'], {'grade': r['grade'], 'reason': r['reason'], 'case': False,
                                            'old_labels': {str(r['label'])}, 'kws': set(), 'source': 'label',
                                            'md_chars': None, 'pdf_chars': None})
        if r['grade'] is not None and (rec['grade'] is None or r['grade'] < rec['grade']):
            rec['grade'], rec['reason'] = r['grade'], r['reason']         # 사유는 최저 등급 행의 것 (recount_grades.page_record 와 같은 규칙)
        rec['case'] = rec['case'] or r['case']
        rec['kws'].add(r['sheet'])
    return pages


def resegment_book(rows, lines, pages_text, prefer_markers=False, stats=None):
    """한 교재: (정렬 또는 마커) → 행 배치 → 쪽별 등급.

    prefer_markers 면 마크다운의 마커로 줄→쪽을 만들고(마커가 빠진 쪽은 hybrid_pages 가 DP 로 쪼갠다) 정렬은
    검증용으로만 계산한다 — 쪽 단위 마커를 이미 가진 교재(변환기가 쪽마다 찍은 것)는 정렬보다 정확하다. 호출부가 마커 밀도로 정한다.
    stats 는 match_rows 의 약한 배정 집계를 받을 dict (선택) — 마커 교재에서는 hybrid_lines(마커 결손 보정으로 쪽이 바뀐 줄 수)도 여기 실린다;
    호출부(main)가 그것을 per_book.hybrid_lines 로 옮기고 match_stats 는 overflow·ambiguous·partial 셋만 남긴다.
    쪽 레코드의 source: text(매칭 행이 있는 쪽) / text-fallback(매칭 행은 없고 미매칭 행의 구 라벨로만 왔지만 본문이 있어
    본문으로 채점) / label(본문도 없어 행 등급의 최저). md_chars·pdf_chars 는 쪽에 붙은 마크다운·PDF 본문의 정규화
    길이(norm_text) — 마커가 빠져 여러 쪽이 한 쪽에 뭉친 곳(md_chars ≫ pdf_chars)이 산출물에서 보이게 한다.
    미매칭 행의 구 라벨이 매칭 행을 이미 가진 쪽과 같은 번호면 그 행은 fallback_rows 로만 세고 kws·case 에는 넣지 않는다 —
    구 라벨은 블록 라벨이라 그 행이 정말 그 쪽에 있는지 모르기 때문이다(키워드·사고사례가 위치 불명 행에서 켜지지 않게).
    반환 (pages, moved, unmatched, line_pages, assigned, idx) 또는 None(줄을 하나도 못 놓음).
    """
    assigned = align_lines(lines, pages_text)
    if prefer_markers and any(MARKER_RE.search(x) for x in lines):
        line_pages = marker_pages(lines)
        if assigned:                                     # 마커 결손 쪽은 DP 로 쪼갠다 (hybrid_pages)
            fixed = hybrid_pages(lines, line_pages, propagate(len(lines), assigned), len(pages_text))
            if stats is not None:                        # 쪽이 바뀐 본문 줄 수 (빈 줄·공백 줄 제외)
                stats['hybrid_lines'] = sum(1 for a, b, ln in zip(line_pages, fixed, lines) if a != b and norm_text(ln))
            line_pages = fixed
    elif assigned:
        line_pages = propagate(len(lines), assigned)
    else:
        return None
    texts = page_texts(lines, line_pages)
    idx = match_rows(rows, lines, stats)
    pages, moved, unmatched = {}, 0, 0

    def pdf_chars(page):                                 # 정규화 길이 — 마크다운 기호·공백을 뺀 실제 글자 수라 md_chars 와 비교된다
        return len(norm_text(pages_text[page - 1])) if isinstance(page, int) and 1 <= page <= len(pages_text) else None

    for r, li in zip(rows, idx):
        matched = li is not None and line_pages[li] is not None
        if not matched:
            unmatched += 1
            page = r['label']
            if page is None:
                continue
        else:
            page = line_pages[li]
            if r['label'] is not None and page != r['label']:
                moved += 1
        rec = pages.get(page)
        if rec is None:
            if page in texts:
                g, _, _, reason = regrade_page(texts[page])
                rec = pages[page] = {'grade': g, 'reason': reason, 'case': False, 'old_labels': set(), 'kws': set(), 'source': 'text',
                                     'matched': 0, 'fallback_rows': 0, 'md_chars': len(norm_text(texts[page])), 'pdf_chars': pdf_chars(page)}
            else:                                            # 본문이 없는 쪽 = 미매칭 행이 구 라벨로 온 경우. 등급은 행 최저(unresolved_pages 와 같은 규칙)
                rec = pages[page] = {'grade': r['grade'], 'reason': r['reason'], 'case': False, 'old_labels': set(), 'kws': set(), 'source': 'label',
                                     'matched': 0, 'fallback_rows': 0, 'md_chars': 0, 'pdf_chars': pdf_chars(page)}
            rec['_fb_kws'], rec['_fb_case'] = set(), False
        elif rec['source'] == 'label' and r['grade'] is not None and (rec['grade'] is None or r['grade'] < rec['grade']):
            rec['grade'], rec['reason'] = r['grade'], r['reason']         # 사유는 최저 등급 행의 것
        if matched:
            rec['matched'] += 1
            rec['case'] = rec['case'] or r['case']
            rec['kws'].add(r['sheet'])
        else:                                                # 위치 불명 행: 일단 따로 모은다
            rec['fallback_rows'] += 1
            rec['_fb_kws'].add(r['sheet'])
            rec['_fb_case'] = rec['_fb_case'] or r['case']
        if r['label'] is not None:
            rec['old_labels'].add(str(r['label']))
    for rec in pages.values():
        fb_kws, fb_case = rec.pop('_fb_kws'), rec.pop('_fb_case')
        if not rec['matched']:                               # 매칭 행이 없는 쪽은 폴백 행이 곧 그 쪽의 전부 — 키워드·사고사례를 그대로 받는다
            rec['kws'] |= fb_kws
            rec['case'] = rec['case'] or fb_case
            if rec['source'] == 'text':                      # 본문은 있지만 그 쪽에 놓인 매칭 행이 없다 — 구 라벨이 데려온 쪽
                rec['source'] = 'text-fallback'
    return pages, moved, unmatched, line_pages, assigned, idx


def sha256_file(path):
    h = hashlib.sha256()
    with open(path, 'rb') as f:
        for chunk in iter(lambda: f.read(1 << 20), b''):
            h.update(chunk)
    return h.hexdigest()


def public_path(p, here=None, home=None):
    """추적 파일(summary.json)에 싣는 경로. 저장소 안이면 상대 경로, 홈 아래면 `~/…`, 그 밖은 마지막 이름만.

    문자열 접두 비교가 아니라 realpath 로 견주므로 심볼릭 링크·`/private/var` 같은 별칭도 걸린다.
    어느 쪽에도 안 걸리는 절대 경로(외장 볼륨, CI 홈)는 그대로 실리지 않는다 — 공개 저장소다.
    """
    rp = os.path.realpath(p)
    for base, prefix in ((os.path.realpath(HERE if here is None else here), ''),
                         (os.path.realpath(os.path.expanduser('~') if home is None else home), '~/')):
        if rp == base:
            return prefix.rstrip('/') or '.'
        if rp.startswith(base.rstrip(os.sep) + os.sep):
            return prefix + os.path.relpath(rp, base)
    return os.path.basename(rp.rstrip(os.sep)) or rp


def write_outputs(books, summary, out_dir):
    os.makedirs(out_dir, exist_ok=True)
    csv_path = os.path.join(out_dir, 'ncs_pages_reseg.csv')

    def write_csv(f):
        w = csv.writer(f)
        w.writerow(['영역', '교재', '페이지', '등급', '등급명', '사고사례', '등급사유', '상태', '출처', 'md자수', 'pdf자수', '구라벨'])
        for name in sorted(books):
            b = books[name]
            for pg in sorted(b['pages']):
                rec = b['pages'][pg]
                w.writerow([b['area'], name, pg, rec['grade'], GRADE_LABEL.get(rec['grade'], ''),
                            '예' if rec['case'] else '아니오', rec['reason'], b['status'], rec.get('source', 'text'),
                            '' if rec.get('md_chars') is None else rec['md_chars'], '' if rec.get('pdf_chars') is None else rec['pdf_chars'],
                            ';'.join(sorted(rec['old_labels'], key=lambda x: int(x) if x.isdigit() else 0))])
    write_atomic(csv_path, write_csv, newline='', encoding='utf-8-sig')        # 파일마다 원자적 교체 (CSV·JSON 쌍 전체가 원자적이지는 않다)
    json_path = os.path.join(out_dir, 'reseg_summary.json')
    write_atomic(json_path, lambda f: json.dump(summary, f, ensure_ascii=False, indent=1))
    return csv_path, json_path


def main():
    ap = argparse.ArgumentParser(description='NCS 워크북 페이지 라벨을 PDF 실제 쪽으로 재세그먼트해 재집계')
    ap.add_argument('--pdf-root', default=os.environ.get('NCS_PDF_ROOT'), help='원본 PDF 루트 (기본 $NCS_PDF_ROOT)')
    ap.add_argument('--md-root', default=DEFAULT_MD_ROOT)
    ap.add_argument('--workbook', default=DEFAULT_WORKBOOK)
    ap.add_argument('--out', default=DEFAULT_OUT)
    ap.add_argument('--paged-dir', default=DEFAULT_PAGED, help='줄→쪽 대응 JSON 을 남길 곳 (gitignore)')
    ap.add_argument('--limit', type=int, default=None, help='앞 N권만 (디버그)')
    ap.add_argument('--force', action='store_true', help='EXPECTED 회귀 검사가 어긋나도 쓴다 (입력이 정당하게 바뀌었을 때만)')
    args = ap.parse_args()
    if args.limit and (os.path.realpath(args.out) == os.path.realpath(DEFAULT_OUT) or os.path.realpath(args.paged_dir) == os.path.realpath(DEFAULT_PAGED)):
        sys.exit('--limit 는 부분 실행이라 추적 산출물(%s)과 기본 대응표 디렉터리(%s)를 덮어쓰지 않습니다 — --out 과 --paged-dir 로 다른 경로를 주십시오' % (DEFAULT_OUT, DEFAULT_PAGED))
    if not os.path.exists(args.workbook):
        sys.exit('워크북이 없습니다: %s' % args.workbook)
    if not args.pdf_root or not os.path.isdir(args.pdf_root):
        sys.exit('PDF 루트가 없습니다: %s — --pdf-root 또는 NCS_PDF_ROOT 로 지정하십시오' % args.pdf_root)
    if not os.path.isdir(args.md_root):
        sys.exit('마크다운 루트가 없습니다: %s' % args.md_root)
    try:
        import fitz  # PyMuPDF
    except ImportError:
        sys.exit('PyMuPDF 가 필요합니다: pip install pymupdf')

    rows = load_rows(args.workbook)
    by_book = collections.OrderedDict()
    for r in rows:
        by_book.setdefault(r['filename'], []).append(r)
    md_index, pdf_index = index_files(args.md_root, '*.md'), index_files(args.pdf_root, '*.pdf')
    print('워크북 %d행, 교재 %d권, 마크다운 %d코드, PDF %d코드' % (len(rows), len(by_book), len(md_index), len(pdf_index)))

    books, per_book, row_map, paged_out, md_paths = {}, {}, [], {}, []   # 대응표는 EXPECTED 검사를 지난 뒤에 쓴다 — 거부된 실행이 추적본과 어긋난 대응표를 남기지 않게

    def mark_unresolved(n, name, brows, area, old_labels, why, label):
        """마크다운·PDF 가 없거나 정렬이 실패한 교재: 구 라벨·구 등급을 그대로 싣고 unresolved 로 센다."""
        books[name] = {'area': area, 'status': 'unresolved', 'rows': len(brows), 'moved_rows': 0,
                       'unmatched_rows': 0, 'old_labels': old_labels, 'align': None, 'pages': unresolved_pages(brows)}
        per_book[name] = {'status': 'unresolved', 'why': why, 'rows': len(brows),
                          'old_pages': len(old_labels), 'new_pages': len(books[name]['pages'])}
        print('  [%2d/%d] %-52s 미해결 (%s)' % (n + 1, len(by_book), name[:52], label))

    for n, (name, brows) in enumerate(by_book.items()):
        if args.limit and n >= args.limit:
            break
        m = CODE_RE.search(name)
        code = m.group(1) if m else None
        md = pick_md(code, name, md_index) if code else None
        pdfs = pdf_index.get(code) if code else None
        area = brows[0]['area']
        old_labels = {str(r['label']) for r in brows if r['label'] is not None}
        if not md or not pdfs:
            why = 'no md' if not md else 'no pdf'
            mark_unresolved(n, name, brows, area, old_labels, why, why)
            continue
        md_paths.append(md)
        with open(md, encoding='utf-8') as f:
            lines = f.read().split('\n')
        doc = fitz.open(pdfs[0])
        try:
            pages_text = [doc[i].get_text() for i in range(len(doc))]
        finally:
            doc.close()
        n_markers = sum(1 for l in lines if MARKER_RE.search(l))
        dense = n_markers >= DENSE_MARKER_RATIO * len(pages_text)
        match = {}
        res = resegment_book(brows, lines, pages_text, prefer_markers=dense, stats=match)
        if res is None:
            mark_unresolved(n, name, brows, area, old_labels, 'alignment failed', '정렬 실패')
            continue
        pages, moved, unmatched, line_pages, assigned, idx = res
        hybrid = match.pop('hybrid_lines', 0)                 # 줄→쪽 보정 수치는 행 매칭 통계와 따로 싣는다
        emptied = emptied_marker_pages(lines, line_pages) if dense else 0   # 마커가 찍힌 쪽인데 본문 줄이 하나도 안 남은 쪽 (보정이 마커를 지웠는지 드러낸다)
        for r, li in zip(brows, idx):
            row_map.append([name, r['sheet'], r['label'], line_pages[li] if li is not None else '', r['grade'], '예' if r['case'] else '아니오'])
        align = None
        if dense:                                         # 마커 교재: 정렬을 마커와 대조 — 후보 줄 기준(lines)과 전체 본문 줄 기준(all_lines) 둘 다
            align = check_alignment(lines, assigned)
            align.update(check_alignment_all(lines, propagate(len(lines), assigned)))
            gaps = gap_lines(lines, len(pages_text))       # 결손 구간은 정답(마커 전파값)이 정의상 틀리므로 따로 뺀 값도 기록
            ng = check_alignment(lines, assigned, exclude=gaps)
            align.update({'nogap_lines': ng['lines'], 'nogap_exact': ng['exact'], 'nogap_near': ng['near']})
        books[name] = {'area': area, 'status': 'resolved', 'rows': len(brows), 'moved_rows': moved,
                       'unmatched_rows': unmatched, 'old_labels': old_labels, 'align': align, 'pages': pages, 'match': match, 'hybrid_lines': hybrid,
                       'emptied_marker_pages': emptied}
        per_book[name] = {'status': 'resolved', 'method': 'markers' if dense else 'alignment', 'rows': len(brows), 'old_pages': len(old_labels),
                          'new_pages': len(pages), 'moved_rows': moved, 'unmatched_rows': unmatched,
                          'pdf_pages': len(pages_text), 'md_markers': n_markers, 'aligned_lines': len(assigned),
                          'align': align, 'match_stats': match, 'hybrid_lines': hybrid, 'emptied_marker_pages': emptied}
        paged_out[code] = {'md': os.path.basename(md), 'pdf': os.path.basename(pdfs[0]), 'line_pages': line_pages}
        print('  [%2d/%d] %-52s 행 %4d  라벨 %3d → 쪽 %3d  이동 %4d  미매칭 %3d  %s%s' % (
            n + 1, len(by_book), name[:52], len(brows), len(old_labels), len(pages), moved, unmatched,
            '마커' if dense else '정렬', ('  정렬검증 %d/%d' % (align['exact'], align['lines'])) if align else ''))

    summary = aggregate(books)
    for name, b in books.items():                     # 교재별 등급 분포 (설계 §3)
        per_book[name]['page_g'] = {g: sum(1 for rec in b['pages'].values() if str(rec['grade']) == g) for g in ('1', '2', '3')}
    summary['per_book'] = per_book
    summary['method_books'] = dict(collections.Counter(v.get('method', v['status']) for v in per_book.values()))
    summary['case_pages'] = [{'book': name, 'page': pg, 'old_labels': sorted(rec['old_labels']), 'grade': rec['grade']}
                             for name, b in books.items() for pg, rec in b['pages'].items() if rec['case']]
    md_used = sorted((os.path.basename(m), sha256_file(m)) for m in md_paths)
    summary['meta'] = {'workbook': os.path.basename(args.workbook), 'workbook_sha256': sha256_file(args.workbook),
                       'md_corpus_sha256': hashlib.sha256('\n'.join('%s\t%s' % x for x in md_used).encode('utf-8')).hexdigest(),   # 쓴 마크다운 84개의 (이름, sha256)
                       'python': sys.version.split()[0], 'pymupdf': getattr(fitz, '__version__', None) or getattr(fitz, 'VersionBind', None),
                       'pdf_root': public_path(args.pdf_root), 'md_root': public_path(args.md_root), 'rows': len(rows),
                       'md_files': sum(len(v) for v in md_index.values()), 'pdf_files': sum(len(v) for v in pdf_index.values()),
                       'limit': args.limit,          # limit 이 있으면 부분 실행 — 전체 실행과 구분한다
                       'rule': 'regrade.grade_page baseline (%s)' % ', '.join('%s=%s' % kv for kv in BASELINE_KW.items()),
                       'run_at': datetime.datetime.now(datetime.timezone.utc).isoformat(timespec='seconds')}
    bad = check_expected(summary) if not args.limit else []
    summary['meta']['expected'] = EXPECTED if not bad and not args.limit else None   # 어긋난 채(--force) 쓰거나 부분 실행이면 기대값을 싣지 않는다
    summary['meta']['expected_mismatch'] = bad or None
    if bad and not args.force:
        sys.exit('EXPECTED 회귀 검사 불일치 — 산출물을 쓰지 않습니다. 입력이 정당하게 바뀌었으면 --force 로 쓰고 EXPECTED 를 갱신하십시오:\n  ' + '\n  '.join(bad))
    if bad:
        print('주의: EXPECTED 와 어긋남 (--force 로 씀): ' + '; '.join(bad))
    os.makedirs(args.paged_dir, exist_ok=True)
    for code, pj in paged_out.items():
        write_atomic(os.path.join(args.paged_dir, code + '.pages.json'), lambda f, pj=pj: json.dump(pj, f))

    def write_rows_map(f):
        w = csv.writer(f)
        w.writerow(['교재', '시트', '구라벨', '새쪽', '구등급', '사고사례'])
        w.writerows(row_map)
    write_atomic(os.path.join(args.paged_dir, 'rows_map.csv'), write_rows_map, newline='', encoding='utf-8-sig')
    csv_path, json_path = write_outputs(books, summary, args.out)
    ac = summary['alignment_check']['overall']
    print('\n페이지 %d (교재 %d) — 등급 %s | 사고사례 %d쪽/%d권 | 이동 행 %d, 미매칭 행 %d | 미해결 %s' % (
        summary['pages'], summary['books'], summary['page_g'], summary['cases_pages'], summary['cases_books'],
        summary['moved_rows'], summary['unmatched_rows'], summary['unresolved']))
    if ac['lines']:
        print('정렬 검증 (쪽 단위 마커 보유 %d권): 후보 줄 정확 %.1f%%, ±1쪽 %.1f%% (%d줄); 전체 본문 줄 정확 %.1f%%, ±1쪽 %.1f%% (%d줄)' % (
            summary['alignment_check']['books'], 100.0 * ac['exact'] / ac['lines'], 100.0 * ac['near'] / ac['lines'], ac['lines'],
            100.0 * ac['all_exact'] / max(ac['all_lines'], 1), 100.0 * ac['all_near'] / max(ac['all_lines'], 1), ac['all_lines']))
    print('→ %s\n→ %s' % (csv_path, json_path))


if __name__ == '__main__':
    main()
